from typing import Optional, List
from sqlalchemy.orm import selectinload, joinedload
from sqlalchemy import or_, func, CompoundSelect
from sqlalchemy.ext.asyncio import AsyncSession
from app.database.models import async_session, User, UserRole, Student, Faculty, Department, Course, Group, Schedule, \
    AttendanceHistoryStudent
from sqlalchemy.exc import SQLAlchemyError
from bot_instance import bot
from sqlalchemy import select, delete
from datetime import datetime
from sqlalchemy import update
import random
from sqlalchemy import or_
from sqlalchemy.orm import aliased
import json
from openpyxl import Workbook
from sqlalchemy import select, func
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError


async def check_admin(telegram_id: str) -> bool:
    try:
        async with async_session() as session:
            async with session.begin():
                result = await session.execute(
                    select(User).where(
                        User.login == telegram_id,
                        or_(User.role == UserRole.ADMIN, User.role == UserRole.SUPERADMIN)
                    )
                )
                admin = result.scalar_one_or_none()

                if admin:
                    return True
                else:
                    return False
    except Exception as e:
        print(f"Error occurred while checking if user is admin: {e}")
        return False

async def get_name(telegram_id: str) -> Optional[str]:
    async with async_session() as session:
        result = await session.execute(
            select(User.full_name).where(telegram_id == User.login)
        )
        name = result.scalar_one_or_none()

        return name

# Add a new user
async def add_user(full_name: str, role: UserRole, login: str) -> bool:
    async with async_session() as session:
        async with session.begin():
            # Проверяем, существует ли уже пользователь с таким логином
            existing_user = await session.execute(
                select(User).where(User.login == login)
            )
            existing_user = existing_user.scalars().first()

            if existing_user:
                return False  # Пользователь с таким логином уже существует

            # Добавляем нового пользователя
            new_user = User(
                full_name=full_name,
                role=role,
                login=login
            )
            session.add(new_user)
            await session.commit()

            return True  # Возвращаем True, если пользователь был успешно добавлен

async def check_super_admin(telegram_id: str) -> bool:
    try:
        async with async_session() as session:
            async with session.begin():
                result = await session.execute(
                    select(User).where(
                        User.login == telegram_id, User.role == UserRole.SUPERADMIN
                    )
                )
                admin = result.scalar_one_or_none()

                if admin:
                    return True
                else:
                    return False
    except Exception as e:
        print(f"Error occurred while checking if user is admin: {e}")
        return False

async def delete_admin(telegram_id: str) -> bool:
    try:
        async with async_session() as session:
            async with session.begin():
                # Проверяем, существует ли админ с таким telegram_id
                result = await session.execute(
                    select(User).where(
                        User.login == telegram_id,
                        or_(User.role == UserRole.ADMIN, User.role == UserRole.SUPERADMIN)
                    )
                )
                admin = result.scalar_one_or_none()

                # Если админ найден, удаляем его
                if admin:
                    await session.execute(
                        delete(User).where(
                            User.login == telegram_id  # Удаляем по telegram_id
                        )
                    )
                    await session.commit()
                    return True
                else:
                    return False
    except Exception as e:
        print(f"Error occurred while deleting admin: {e}")
        return False

async def delete_teacher(telegram_id: str) -> bool:
    try:
        async with async_session() as session:
            async with session.begin():
                result = await session.execute(
                    select(User).where(
                        User.login == telegram_id, User.role == UserRole.TEACHER)
                    )
                teacher = result.scalar_one_or_none()

                if teacher:
                    await session.execute(
                        delete(User).where(
                            User.login == telegram_id  # Удаляем по telegram_id
                        )
                    )
                    await session.commit()
                    return True
                else:
                    return False
    except Exception as e:
        print(f"Error occurred while deleting admin: {e}")
        return False

async def export_users_to_excel_by_role(role: UserRole, filename: str) -> bool:
    try:
        # Открываем сессию базы данных
        async with async_session() as session:
            # Выполняем запрос на получение пользователей с указанной ролью
            if role == UserRole.TEACHER:
                result = await session.execute(
                    select(User).where(User.role == role)
                )
                users = result.scalars().all()
            else:
                result = await session.execute(
                    select(User).where(User.role != UserRole.TEACHER)
                )
                users = result.scalars().all()

            if not users:
                print("No users found for the given role.")
                return False

            # Создаём Excel файл
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = f"{role.value} Users"

            # Записываем заголовки
            headers = ["User ID", "Full Name", "Role", "Login", "Created At"]
            sheet.append(headers)

            # Заполняем данные
            for user in users:
                sheet.append([
                    user.user_id,
                    user.full_name,
                    user.role.value,
                    user.login,
                    user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else ""
                ])

            # Сохраняем файл
            workbook.save(filename)
            return True
    except Exception as e:
        print(f"Error in export_users_to_excel_by_role: {e}")
        return False



async def get_university_statistics() -> str:
    try:
        async with async_session() as session:
            # Общее количество пользователей
            total_users = await session.scalar(select(func.count(User.user_id)))
            total_admins = await session.scalar(
                select(func.count(User.user_id)).where(User.role == UserRole.ADMIN)
            )
            total_super_admins = await session.scalar(
                select(func.count(User.user_id)).where(User.role == UserRole.SUPERADMIN)
            )
            total_admins = total_admins + total_super_admins
            total_teachers = await session.scalar(
                select(func.count(User.user_id)).where(User.role == UserRole.TEACHER)
            )
            total_students = await session.scalar(select(func.count(Student.student_id)))

            # Количество факультетов и кафедр
            total_faculties = await session.scalar(select(func.count(Faculty.faculty_id)))
            total_departments = await session.scalar(select(func.count(Department.department_id)))

            # Количество курсов и групп
            total_courses = await session.scalar(select(func.count(Course.course_id)))
            total_groups = await session.scalar(select(func.count(Group.group_id)))

            # Количество записей в расписании
            total_schedule_entries = await session.scalar(select(func.count(Schedule.schedule_id)))

            # Пропущенные занятия
            total_absences = await session.scalar(
                select(func.count(AttendanceHistoryStudent.id)).where(
                    AttendanceHistoryStudent.status == False
                )
            )

            # Формирование текста статистики
            stats_text = (
                f"📊 *Университетская статистика*\n\n"
                f"👥 Пользователи:\n"
                f"   - Всего: {total_users}\n"
                f"   - Администраторы: {total_admins}\n"
                f"   - Преподаватели: {total_teachers}\n"
                f"   - Студенты: {total_students}\n\n"
                f"🏢 Структура:\n"
                f"   - Факультеты: {total_faculties}\n"
                f"   - Кафедры: {total_departments}\n"
                f"   - Курсы: {total_courses}\n"
                f"   - Группы: {total_groups}\n\n"
                f"📅 Расписание:\n"
                f"   - Записей: {total_schedule_entries}\n\n"
                f"❌ Пропущенные занятия:\n"
                f"   - Всего пропусков: {total_absences}\n"
            )

            return stats_text
    except Exception as e:
        print(f"Error occurred while fetching statistics: {e}")
        return "Произошла ошибка при получении статистики. Попробуйте позже."

async def export_faculties_to_excel(filename: str) -> bool:
    try:
        # Открываем сессию базы данных
        async with async_session() as session:
            # Выполняем запрос на получение всех факультетов
            result = await session.execute(select(Faculty))
            faculties = result.scalars().all()

            if not faculties:
                print("No faculties found.")
                return False

            # Создаём Excel файл
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Faculties"

            # Записываем заголовки
            headers = ["Faculty ID", "Name"]
            sheet.append(headers)

            # Заполняем данные
            for faculty in faculties:
                sheet.append([faculty.faculty_id, faculty.name])

            # Сохраняем файл
            workbook.save(filename)
            return True
    except Exception as e:
        print(f"Error in export_faculties_to_excel: {e}")
        return False



async def import_faculties_from_excel(filename: str) -> str:
    try:
        # Загружаем файл Excel
        workbook = load_workbook(filename=filename)
        sheet = workbook.active  # Предполагается, что данные на первом листе

        # Проверяем наличие заголовков
        headers = ["Faculty ID", "Name"]
        if not all(header == sheet.cell(row=1, column=i + 1).value for i, header in enumerate(headers)):
            return "Неверный формат Excel файла. Проверьте заголовки: Faculty ID, Name."

        # Открываем сессию для записи в БД
        async with async_session() as session:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                faculty_id, name = row

                if not name:  # Пропускаем записи с пустыми названиями
                    continue

                # Проверяем, существует ли запись с таким ID (если `faculty_id` передаётся)
                if faculty_id:
                    faculty = await session.get(Faculty, faculty_id)
                    if faculty:
                        faculty.name = name  # Обновляем название
                    else:
                        # Добавляем новую запись
                        new_faculty = Faculty(faculty_id=faculty_id, name=name)
                        session.add(new_faculty)
                else:
                    # Добавляем новую запись без `faculty_id`
                    new_faculty = Faculty(name=name)
                    session.add(new_faculty)

            await session.commit()
        return "Данные успешно импортированы в БД."
    except FileNotFoundError:
        return "Файл не найден. Проверьте путь."
    except IntegrityError as e:
        return f"Ошибка целостности данных: {e.orig}"
    except Exception as e:
        return f"Произошла ошибка при импорте: {e}"


async def export_departments_to_excel(filename: str) -> bool:
    try:
        # Открываем сессию базы данных
        async with async_session() as session:
            # Выполняем запрос на получение всех кафедр
            result = await session.execute(select(Department))
            departments = result.scalars().all()

            if not departments:
                print("No departments found.")
                return False

            # Создаём Excel файл
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Departments"

            # Записываем заголовки
            headers = ["Department ID", "Name", "Faculty ID"]
            sheet.append(headers)

            # Заполняем данные
            for department in departments:
                sheet.append([department.department_id, department.name, department.faculty_id])

            # Сохраняем файл
            workbook.save(filename)
            return True
    except Exception as e:
        print(f"Error in export_departments_to_excel: {e}")
        return False


async def import_departments_from_excel(filename: str) -> str:
    try:
        # Загружаем файл Excel
        workbook = load_workbook(filename=filename)
        sheet = workbook.active

        # Проверяем наличие заголовков
        headers = ["Department ID", "Name", "Faculty ID"]
        if not all(header == sheet.cell(row=1, column=i + 1).value for i, header in enumerate(headers)):
            return "Неверный формат Excel файла. Проверьте заголовки: Department ID, Name, Faculty ID."

        async with async_session() as session:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                department_id, name, faculty_id = row

                if not name or not faculty_id:
                    continue  # Пропускаем записи с пустыми значениями

                # Проверяем, существует ли запись с таким ID (если `department_id` передаётся)
                if department_id:
                    department = await session.get(Department, department_id)
                    if department:
                        department.name = name
                        department.faculty_id = faculty_id  # Обновляем faculty_id
                    else:
                        # Добавляем новую запись
                        new_department = Department(department_id=department_id, name=name, faculty_id=faculty_id)
                        session.add(new_department)
                else:
                    # Добавляем новую запись без `department_id`
                    new_department = Department(name=name, faculty_id=faculty_id)
                    session.add(new_department)

            await session.commit()
        return "Данные успешно импортированы в БД."
    except FileNotFoundError:
        return "Файл не найден. Проверьте путь."
    except IntegrityError as e:
        return f"Ошибка целостности данных: {e.orig}"
    except Exception as e:
        return f"Произошла ошибка при импорте: {e}"


